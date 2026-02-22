
"""
BOQ Currency Converter  —  KD → QAR
Run:  python3 app.py
Open: http://localhost:8000
"""

import asyncio, io, json, re, tempfile, uuid
from decimal import Decimal, ROUND_HALF_UP
from pathlib import Path

import fitz
import openpyxl
import pdfplumber
import uvicorn
from fastapi import FastAPI, File, Form, UploadFile
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import FileResponse, HTMLResponse, StreamingResponse
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

JOBS: dict[str, dict] = {}
TMP  = Path(tempfile.gettempdir()) / "boq_jobs"
TMP.mkdir(exist_ok=True)

app = FastAPI()
app.add_middleware(CORSMiddleware, allow_origins=["*"],
                   allow_methods=["*"], allow_headers=["*"])

MONEY_RE = re.compile(r"^\d{1,3}(?:,\d{3})*\.\d{3}$")

def convert_money(text: str, rate: Decimal):
    t = text.strip()
    if t == "KD": return "QAR"
    if MONEY_RE.match(t):
        v = Decimal(t.replace(",", ""))
        return float((v * rate).quantize(Decimal("0.001"), rounding=ROUND_HALF_UP))
    return text


# ── PDF → PDF ──────────────────────────────────────────────────────────────
async def run_pdf_to_pdf(src: bytes, rate: Decimal, job_id: str):
    MIN_FONT = 5.0
    _FC: dict = {}

    def gf(flags):
        bold, italic = bool(flags & 16), bool(flags & 2)
        name = ("tibi" if italic else "tibo") if bold else ("tiit" if italic else "tiro")
        if name not in _FC: _FC[name] = (fitz.Font(name), name)
        return _FC[name]

    def rgb(c): return (c >> 16) / 255, ((c >> 8) & 0xFF) / 255, (c & 0xFF) / 255

    def place(page, text, x0, y, x1, fo, fn, fs, color, right_align):
        cw = x1 - x0
        nw = fo.text_length(text, fontsize=fs)
        if nw > cw + 0.5:
            fs = max(MIN_FONT, fs * (cw + 0.5) / nw)
            nw = fo.text_length(text, fontsize=fs)
        x = (x1 - nw) if right_align else x0
        page.insert_text((x, y), text, fontname=fn, fontsize=fs, color=color)

    doc      = fitz.open(stream=src, filetype="pdf")
    N        = len(doc)
    replaced = 0

    for pn, page in enumerate(doc):
        reps = []
        for block in page.get_text("dict")["blocks"]:
            for line in block.get("lines", []):
                for span in line.get("spans", []):
                    clean = span["text"].strip()
                    if   clean == "KD":        new = "QAR"
                    elif MONEY_RE.match(clean):
                        v = convert_money(clean, rate)
                        new = f"{v:,.3f}" if isinstance(v, float) else v
                    else: continue
                    bbox = fitz.Rect(span["bbox"])
                    fo, fn = gf(span["flags"])
                    fs, o  = span["size"], span["origin"]
                    ow     = fo.text_length(clean, fontsize=fs)
                    right  = abs(bbox.x1 - (o[0] + ow)) < abs(o[0] - bbox.x0)
                    reps.append((bbox, o, new, fo, fn, fs, rgb(span["color"]), right))

        if reps:
            for bbox, *_ in reps:
                page.add_redact_annot(bbox + (-1, -1, 2, 1), fill=(1, 1, 1))
            page.apply_redactions(images=fitz.PDF_REDACT_IMAGE_NONE)
            for bbox, o, new, fo, fn, fs, color, right in reps:
                place(page, new, bbox.x0, o[1], bbox.x1, fo, fn, fs, color, right)
            replaced += len(reps)

        JOBS[job_id].update({"progress": round((pn+1)/N*100),
                              "page": pn+1, "pages": N, "replaced": replaced})
        await asyncio.sleep(0)

    out = TMP / f"{job_id}.pdf"
    doc.save(str(out), garbage=4, deflate=True, clean=True)
    doc.close()
    JOBS[job_id].update({"status": "done", "file": str(out),
                          "replaced": replaced, "pages": N})


# ── PDF → XLSX ─────────────────────────────────────────────────────────────
async def run_pdf_to_xlsx(src: bytes, rate: Decimal, job_id: str):
    THIN   = Side(style="thin", color="FF000000")
    thin_b = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
    COL_W  = [7, 45, 8, 7, 10, 16, 12, 10, 10, 16]
    FMT    = "#,##0.000"

    def to_num(t):
        t = t.strip()
        if re.match(r"^\d{1,3}(?:,\d{3})*\.\d+$", t) or re.match(r"^\d+\.\d+$", t):
            return float(t.replace(",", ""))
        if re.match(r"^\d+$", t): return int(t)
        return None

    def expand(table):
        out = []
        for row in table:
            if row[0] and "\n" in str(row[0]):
                cs = [str(c).split("\n") if c else [] for c in row]
                for i in range(max(len(c) for c in cs)):
                    out.append([c[i] if i < len(c) else "" for c in cs])
            else:
                out.append([str(c) if c else "" for c in row])
        return out

    def style_row(ws, r, is_hdr=False, is_tot=False):
        bg = "FFB2EBF2" if is_hdr else ("FFD0D0D0" if is_tot else "FFFFFFFF")
        for ci in range(1, 11):
            c = ws.cell(r, ci)
            c.font   = Font(name="Calibri", bold=(is_hdr or is_tot), size=9)
            c.fill   = PatternFill("solid", start_color=bg)
            c.border = thin_b
            c.alignment = Alignment(
                horizontal=("center" if ci == 1 or is_hdr else
                            "right"  if ci in (3, 5, 6, 7, 8, 9, 10) else "left"),
                vertical="center", wrap_text=(ci == 2))
            if ci in (5, 6, 7, 8, 9, 10) and isinstance(c.value, float):
                c.number_format = FMT

    # احفظ كملف مؤقت بدل BytesIO — أكثر استقراراً
    tmp_in = TMP / f"{job_id}_in.pdf"
    tmp_in.write_bytes(src)

    wb       = openpyxl.Workbook()
    wb.remove(wb.active)
    replaced = 0
    TS       = {"vertical_strategy": "lines", "horizontal_strategy": "lines"}

    try:
        with pdfplumber.open(str(tmp_in)) as pdf:
            N = len(pdf.pages)
            for pi, page in enumerate(pdf.pages):
                tables = page.extract_tables(TS)
                if tables:
                    ws = wb.create_sheet(f"P{pi+1:03d}")
                    for i, w in enumerate(COL_W, 1):
                        ws.column_dimensions[get_column_letter(i)].width = w
                    er = 1
                    for table in tables:
                        for row in expand(table):
                            item   = row[0].strip() if row[0] else ""
                            is_hdr = any(h in str(row) for h in ("DESCRIPTION", "Work Done"))
                            is_tot = item in ("TT","TG") or bool(re.match(r"^T\d+$", item))
                            for ci, raw in enumerate(row, 1):
                                cv = convert_money(raw, rate) if raw else ""
                                if isinstance(cv, float): replaced += 1
                                v  = cv if isinstance(cv, float) else (to_num(str(cv)) or (cv if cv else None))
                                ws.cell(er, ci, v)
                            style_row(ws, er, is_hdr, is_tot)
                            ws.row_dimensions[er].height = 15
                            er += 1
                        er += 1

                JOBS[job_id].update({"progress": round((pi+1)/N*100),
                                      "page": pi+1, "pages": N,
                                      "replaced": replaced})
                await asyncio.sleep(0)
    finally:
        tmp_in.unlink(missing_ok=True)

    out = TMP / f"{job_id}.xlsx"
    wb.save(str(out))
    JOBS[job_id].update({"status": "done", "file": str(out),
                          "pages": N, "replaced": replaced})


# ── API ────────────────────────────────────────────────────────────────────
@app.post("/upload")
async def upload(file: UploadFile = File(...),
                 mode: str  = Form("pdf"),
                 rate: str  = Form("11.9")):
    job_id = str(uuid.uuid4())
    src    = await file.read()
    ext    = "pdf" if mode == "pdf" else "xlsx"
    JOBS[job_id] = {
        "status":   "running", "progress": 0,
        "page":     0,         "pages":    0,
        "replaced": 0,         "mode":     mode,
        "filename": file.filename,
        "out_name": Path(file.filename).stem + f"_QAR.{ext}",
    }

    async def run():
        try:
            if mode == "pdf":
                await run_pdf_to_pdf(src, Decimal(rate), job_id)
            else:
                await run_pdf_to_xlsx(src, Decimal(rate), job_id)
        except Exception as e:
            JOBS[job_id].update({"status": "error", "error": str(e)})

    asyncio.create_task(run())
    return {"job_id": job_id}


@app.get("/progress/{job_id}")
async def progress_stream(job_id: str):
    async def stream():
        while True:
            job = JOBS.get(job_id, {})
            if not job:
                yield f"data:{json.dumps({'error':'not found'})}\n\n"; break
            yield f"data:{json.dumps(job)}\n\n"
            if job.get("status") in ("done", "error"): break
            await asyncio.sleep(0.25)

    return StreamingResponse(stream(), media_type="text/event-stream",
                             headers={"Cache-Control": "no-cache",
                                      "X-Accel-Buffering": "no"})


@app.get("/download/{job_id}")
async def download(job_id: str):
    job = JOBS.get(job_id, {})
    if not job or job.get("status") != "done":
        return {"error": "not ready"}
    mt = ("application/pdf" if job["out_name"].endswith(".pdf")
          else "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    return FileResponse(job["file"], media_type=mt,
                        headers={"Content-Disposition":
                                 f'attachment; filename="{job["out_name"]}"'})


# ── Frontend ───────────────────────────────────────────────────────────────
@app.get("/", response_class=HTMLResponse)
async def root():
    return open(__file__.replace("app.py", "index.html"), encoding="utf-8").read()


if __name__ == "__main__":
    import os
    print("\n" + "─"*52)
    print("  BOQ Converter  ·  KD → QAR")
    print("─"*52)
    print("  ▶  http://localhost:8000")
    print("  ■  Ctrl+C to stop")
    print("─"*52 + "\n")
    uvicorn.run(app, host="0.0.0.0",
                port=int(os.environ.get("PORT", 8000)),
                log_level="warning")